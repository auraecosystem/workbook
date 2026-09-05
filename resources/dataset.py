import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Path to target Excel workbook
excel_path = "/mnt/data/Aura.xlsx"

# -----------------------------------------------------------------------------
# 1. EXPANDED STEM DATASETS
# -----------------------------------------------------------------------------
data_pure_math = {
    "Topic": [
        "Abstract Algebra",
        "Calculus & Analysis",
        "Number Theory",
        "Topology",
        "Differential Geometry",
    ],
    "Formula/Theorem": [
        "First Isomorphism Theorem: G/Kernel(φ) ≅ Im(φ)",
        "Fundamental Theorem of Calculus: ∫[a,b] f(x)dx = F(b) - F(a)",
        "Prime Number Theorem: π(n) ~ n / ln(n)",
        "Euler Characteristic: χ = V - E + F",
        "Gauss-Bonnet Theorem: ∫∫_M K dA + ∫_∂M k_g ds = 2π χ(M)",
    ],
    "Subfield": [
        "Group Theory",
        "Real Analysis",
        "Analytic Number Theory",
        "Algebraic Topology",
        "Differential Geometry",
    ],
    "Notes": [
        "Relates quotient groups to homomorphic images",
        "Connects differentiation and integration",
        "Describes asymptotic distribution of prime numbers",
        "Topological invariant for polyhedra and surfaces",
        "Links intrinsic geometry to topological structure",
    ],
}

data_further_math = {
    "Concept": [
        "Matrix Multiplication",
        "Eigenspaces",
        "Ordinary Diff Eq",
        "Complex Analysis",
        "Fourier Series",
    ],
    "Expression": [
        "C_{i,j} = ∑_{k=1}^n A_{i,k} B_{k,j}",
        "det(A - λI) = 0",
        "dy/dx + P(x)y = Q(x)",
        "f(z) = u(x,y) + i v(x,y)",
        "f(x) = a0/2 + ∑ [an cos(nx) + bn sin(nx)]",
    ],
    "Domain": [
        "Linear Algebra",
        "Linear Algebra",
        "Differential Equations",
        "Complex Variables",
        "Harmonic Analysis",
    ],
    "Application": [
        "Transformations & linear systems",
        "Dimensionality reduction & quantum states",
        "Modeling dynamic systems and growth",
        "Fluid dynamics & conformal mapping",
        "Signal processing & wave approximation",
    ],
}

data_applied_physics = {
    "Field": [
        "Classical Mechanics",
        "Thermodynamics",
        "Electromagnetism",
        "Quantum Mechanics",
        "General Relativity",
    ],
    "Equation": [
        "F = d(mv)/dt",
        "dU = δQ - δW",
        "∇ · E = ρ / ε₀",
        "iℏ ∂Ψ/∂t = ĤΨ",
        "G_μν + Λ g_μν = (8πG/c⁴) T_μν",
    ],
    "Governing Law / Model": [
        "Newton's Second Law",
        "First Law of Thermodynamics",
        "Gauss's Law (Maxwell)",
        "Time-Dependent Schrödinger Equation",
        "Einstein Field Equations",
    ],
    "Physical Significance": [
        "Relates force to momentum change",
        "Conservation of energy in closed thermodynamic systems",
        "Electric flux proportional to enclosed charge",
        "Determines time evolution of quantum wavefunctions",
        "Describes gravity as spacetime curvature caused by mass-energy",
    ],
}

data_reasoning_logic = {
    "Premise A": [
        "All humans are mortal",
        "If P then Q",
        "All A are B",
        "P ∨ Q",
        "A → B",
    ],
    "Premise B": [
        "Socrates is human",
        "P is true",
        "All B are C",
        "¬P",
        "¬B",
    ],
    "Rule of Inference": [
        "Categorical Syllogism",
        "Modus Ponens",
        "Hypothetical Syllogism",
        "Disjunctive Syllogism",
        "Modus Tollens",
    ],
    "Conclusion": [
        "Socrates is mortal",
        "Q is true",
        "All A are C",
        "Q is true",
        "¬A",
    ],
    "Truth Value": ["Valid", "Valid", "Valid", "Valid", "Valid"],
}

data_simulation = {
    "Simulation Target": [
        "Projectile Motion",
        "Heat Diffusion",
        "Quantum Tunneling",
        "N-Body Gravitational System",
        "Monte Carlo Integration",
    ],
    "System Parameters & Setup": [
        "v₀ = 20 m/s, θ = 45°, g = 9.81 m/s², air resistance = 0",
        "Rod L = 1.0 m, k = 200 W/m·K, boundary T = 100°C / 0°C",
        "Potential barrier V₀ > E, barrier width d = 0.5 nm",
        "N = 3 celestial bodies, G = 6.674e-11, timestep Δt = 1s",
        "Function f(x) = e^(-x²), range [0, 1], samples N = 10⁶",
    ],
    "Numerical Method": [
        "Runge-Kutta 4th Order (RK4)",
        "Finite Difference Method (FTCS)",
        "Split-Operator Spectral Method",
        "Verlet Integration",
        "Stochastic Uniform Sampling",
    ],
    "Expected Output / Metric": [
        "Theoretical Range R ≈ 40.77 m",
        "1D steady-state linear temperature profile",
        "Transmission coefficient T > 0 through barrier",
        "Stable or chaotic orbits / energy drift < 0.01%",
        "Approximated Integral value ≈ 0.7468",
    ],
}

stem_sheets = {
    "Pure_Mathematics": pd.DataFrame(data_pure_math),
    "Further_Mathematics": pd.DataFrame(data_further_math),
    "Applied_Physics": pd.DataFrame(data_applied_physics),
    "Reasoning_Logic": pd.DataFrame(data_reasoning_logic),
    "Simulation_Problems": pd.DataFrame(data_simulation),
}

# -----------------------------------------------------------------------------
# 2. WRITE DATA & APPLY OPENPYXL STYLING
# -----------------------------------------------------------------------------
with pd.ExcelWriter(
    excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
) as writer:
    for sheet_name, df in stem_sheets.items():
        # Export DataFrame to target sheet
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Access worksheet for custom formatting
        ws = writer.sheets[sheet_name]
        ws.views.sheetView[0].showGridLines = True  # Preserve gridlines

        # Design system styles
        header_fill = PatternFill(
            start_color="1F497D", end_color="1F497D", fill_type="solid"
        )  # Deep Navy
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11, bold=False, color="000000")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell_alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True
        )

        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Format data rows & borders
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 22
            for cell in row:
                cell.font = data_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Enable AutoFilter across all populated columns
        ws.auto_filter.ref = ws.dimensions

        # Dynamically set column widths based on content length
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            # Set width with extra padding, capped at 50 for wrapped text
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 4, 15), 50
            )
